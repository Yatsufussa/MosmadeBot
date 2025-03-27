import logging
from datetime import datetime

from aiogram.utils.markdown import hbold
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill

import buttons.inline_buttons as kb
import pandas as pd
import tempfile
from database.engine import *
from aiogram import F, Router, types, Bot
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, ReplyKeyboardRemove, Message
from aiogram.types.input_file import FSInputFile
from ChatFilter.chat_type import ChatTypeFilter, IsAdmin

from database.orm_queries import orm_add_category, orm_delete_category, \
    orm_get_category_by_id, orm_add_product, orm_get_product_by_id, \
    orm_update_product_price, orm_update_product_photo, orm_delete_product_by_id, \
    orm_update_category_sex, orm_update_category_name_ru, orm_update_category_name_uz, \
    orm_update_product_description_uz, orm_update_product_description_ru, orm_update_product_name_ru, \
    orm_update_product_name_uz, orm_get_all_user_ids, orm_get_all_excel_orders, orm_delete_all_excel_orders, \
    orm_get_user_by_tg_id, update_excel_order_status_to_cancelled, orm_get_order_by_id, orm_delete_order_by_id, \
    update_order_status_to_finished, orm_add_promocode, orm_product_exists, orm_promocode_exists, orm_delete_promocode, \
    orm_delete_bonus_product, orm_add_bonus_product, orm_update_product_video
from language_dictionary.language import GENDER_MAPPING, MESSAGES

admin_router = Router()
admin_router.message.filter(ChatTypeFilter(["private"]), IsAdmin())

GROUP_CHAT_IDS_WITH_THREADS = [
    {"chat_location": -1002408666314, "message_thread_id": 73},  # Example thread ID
]


# region ADMINS CATEGORY MANIPULATION STATES
class AddCategory(StatesGroup):
    add_name_ru = State()
    add_name_uz = State()
    add_sex = State()


class ChangeCategory(StatesGroup):
    ch_name_ru = State()
    ch_name_uz = State()
    ch_sex = State()


class DeleteCategory(StatesGroup):
    waiting_for_confirmation = State()


# endregion

# region ADMINS PRODUCT MANIPULATION STATES
class AddProduct(StatesGroup):
    add_name_ru = State()
    add_name_uz = State()
    add_description_ru = State()
    add_description_uz = State()
    add_price = State()
    add_video = State()  # Only video state is needed


class ChangeProduct(StatesGroup):
    ch_name_ru = State()
    ch_name_uz = State()
    ch_description_ru = State()
    ch_description_uz = State()
    ch_price = State()
    ch_video = State()


class DeleteProduct(StatesGroup):
    delete_product = State()


class AddPromoCode(StatesGroup):
    enter_promo_code = State()  # Step 1: Enter the promo code text
    enter_product_id = State()  # Step 2: Enter the product ID (optional)
    enter_discount = State()  # Step 3: Enter the discount percentage
    enter_expiry_date = State()  # Step 4: Enter the expiry date (optional)


# endregion

class AddBonusProduct(StatesGroup):
    name_ru = State()  # Russian name
    name_uz = State()  # Uzbek name
    description_ru = State()  # Russian description
    description_uz = State()  # Uzbek description
    image_url = State()  # Image URL
    required_referrals = State()  # Number of required referrals


# region ADMIN PANEL MAIN CATALOG PRODUCTS CATEGORIES
@admin_router.message(Command("admin"))
async def admin_features(message: types.Message):
    await message.answer(
        "👨‍💼 <b>Admin Commands Guide</b>:\n\n"
        "📌 /admin - Open Admin Menu\n"
        "🛑 /stop - Stop the process\n"
        "📢 /newsletter [text] - Send a broadcast message\n"
        "✅ /OrderFinish [Order ID] - Mark order as delivered\n"
        "❌ /OrderCancelation [Order ID] - Cancel and delete an order\n"
        "📍 /GetLocation [Order ID] - Get the location of an order\n"
        "📜 /getallorders - Download all orders as an Excel document\n"
        "🎁 /add_promocode - Add a new promo code\n"
        "💰 /add_bonus - Add a bonus\n"
        "🗑️ /delete_bonus - Delete a bonus\n"
        "🚫 /promocodedelete - Delete a promo code\n\n"
        "⚠️ /ExcelOrdersClear - <b>WARNING:</b> Deletes all Excel order records! (Recommended to download first via /getallorders)\n",
        reply_markup=kb.admin_main,
        parse_mode="HTML"
    )


@admin_router.message(Command("stop"))
async def admin_stop(message: types.Message):
    await message.delete()
    await message.answer('Вы в Админской части поздравляю',
                         reply_markup=kb.admin_main)


@admin_router.callback_query(F.data == 'category')
async def admin_category(callback: CallbackQuery):
    await callback.answer('Вы нажали на категорию')
    await callback.message.edit_text('Действие для Категории',
                                     reply_markup=kb.admin_category)


@admin_router.callback_query(F.data == 'product')
async def admin_product(callback: CallbackQuery):
    await callback.answer('Вы нажали на Продукт')
    await callback.message.edit_text('Действие для Товара',
                                     reply_markup=kb.admin_product)


# endregion


# region ADMIN CATEGORY ADD, CHANGE, DELETE Handlers
# region ADD CATEGORY
@admin_router.callback_query(F.data == 'add_category')
async def admin_add_category(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AddCategory.add_name_ru)
    await callback.answer('')
    await callback.message.answer('Введите имя для новой категории', reply_markup=ReplyKeyboardRemove())


@admin_router.message(AddCategory.add_name_ru)
async def admin_add_category_name(message: Message, state=FSMContext):
    category_name_ru = message.text

    # Validate category name length
    if not (2 <= len(category_name_ru) <= 30):
        await message.answer(
            "Имя категории должно быть длиной от 2 до 30 символов. Пожалуйста, введите допустимое имя...")
        return

    await state.update_data(name_ru=category_name_ru)
    await message.answer('Введите Название Категории на Узбекском', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddCategory.add_name_uz)


@admin_router.message(AddCategory.add_name_uz)
async def admin_add_category_name_uz(message: Message, state=FSMContext):
    category_name_uz = message.text

    # Validate category name length
    if not (2 <= len(category_name_uz) <= 30):
        await message.answer(
            "Имя категории должно быть длиной от 2 до 30 символов. Пожалуйста, введите допустимое имя...")
        return

    # Save the Uzbek category name in state
    await state.update_data(name_uz=category_name_uz)

    # Retrieve both names from state
    data = await state.get_data()
    category_data = {
        'name_ru': data['name_ru'],
        'name_uz': data['name_uz']
    }

    try:
        # Add the category to the database
        await orm_add_category(category_data)
        await message.answer("Категория успешно добавлена", reply_markup=kb.admin_category)
    except Exception as e:
        await message.answer(f"Произошла ошибка при добавлении категории: {e}")

    # Clear state after finishing the process
    await state.clear()


# endregion ADD CATEGORY

# region CHANGE CATEGORY NAME
@admin_router.callback_query(F.data == 'change_category')
async def category_operation_type(callback: CallbackQuery, state: FSMContext):
    await state.update_data(sex=None)  # Set default gender to None to allow skipping
    await callback.answer('')

    # Option to skip gender selection and directly go to category selection
    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.categories(
            back_callback='to_admin_category',
            page=1,
            categories_per_page=4,
            language='ru',
            sex=None  # Default to None for all categories
        )
    )


@admin_router.callback_query(F.data.startswith('admingender_'))
async def category_gender_selection(callback: CallbackQuery, state: FSMContext):
    gender = callback.data.split('_')[1]
    sex = GENDER_MAPPING.get(gender, gender)  # Convert gender term to Russian

    await state.update_data(sex=sex)

    # Fetch categories based on gender selection
    await callback.answer('')
    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.categories(
            back_callback='to_admin_category',
            page=1,
            categories_per_page=4,
            language='ru',
            sex=sex
        )
    )


@admin_router.callback_query(F.data.startswith('AdminCategories_'))
async def admin_categories_pagination(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[1])
    data = await state.get_data()
    sex = data.get('sex', None)  # Retrieve selected gender or default to None
    await callback.answer('')
    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.categories(
            back_callback='to_admin_category',
            page=page,
            categories_per_page=4,
            language='ru',
            sex=sex
        )
    )


@admin_router.callback_query(F.data.startswith('AdminCategory_'))
async def select_category(callback: CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split('_')[-1])
    category = await orm_get_category_by_id(category_id)

    if category:
        await state.update_data(category_id=category_id)
        await callback.answer('')
        await callback.message.edit_text(
            text=f'Вы выбрали категорию {category.name_ru} с ID {category_id}.\nЧто изменить?',
            reply_markup=kb.admin_category_change
        )
    else:
        await callback.answer('Категория не найдена.', show_alert=True)


@admin_router.callback_query(F.data == 'change_c_name')
async def start_change_category_name(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ChangeCategory.ch_name_ru)
    await callback.answer('')
    await callback.message.answer('Введите новое имя категории (Ru)', reply_markup=ReplyKeyboardRemove())


@admin_router.message(ChangeCategory.ch_name_ru)
async def set_new_category_name_ru(message: Message, state: FSMContext):
    new_name_ru = message.text.strip()
    # Validate new category name length
    if not (2 <= len(new_name_ru) <= 30):
        await message.answer(
            "Имя категории должно быть длиной от 2 до 30 символов. Пожалуйста, введите допустимое имя.")
        return

    data = await state.get_data()
    category_id = data.get('category_id')

    if not category_id:
        await message.answer('Не удалось получить ID категории. Пожалуйста, попробуйте снова.',
                             reply_markup=kb.admin_category)
        await state.clear()
        return

    updated = await orm_update_category_name_ru(category_id, new_name_ru)

    if updated:
        await message.answer(f'Имя категории обновлено на {new_name_ru}.', reply_markup=ReplyKeyboardRemove())
        await state.set_state(ChangeCategory.ch_name_uz)
        await message.answer('Kategoriya nomini yangilang (Uz)', reply_markup=ReplyKeyboardRemove())
    else:
        await message.answer(f'Категория с ID {category_id} не найдена.', reply_markup=kb.admin_category)
        await state.clear()


@admin_router.message(ChangeCategory.ch_name_uz)
async def set_new_category_name_uz(message: Message, state: FSMContext):
    new_name_uz = message.text.strip()

    # Validate new category name length
    if not (2 <= len(new_name_uz) <= 30):
        await message.answer(
            "Kategoriya nomi 2 dan 30 belgidan iborat bo'lishi kerak. Iltimos, qabul qilingan nomni kiriting.")
        return

    data = await state.get_data()
    category_id = data.get('category_id')

    if not category_id:
        await message.answer('Не удалось получить ID категории. Пожалуйста, попробуйте снова.',
                             reply_markup=kb.admin_category)
        await state.clear()
        return

    updated = await orm_update_category_name_uz(category_id, new_name_uz)

    if updated:
        await message.answer(f'Kategoriya nomi {new_name_uz} ga o`zgartirildi.', reply_markup=kb.admin_category)
    else:
        await message.answer(f'ID {category_id} bo`yicha kategoriya topilmadi.', reply_markup=kb.admin_category)
    await state.clear()


# endregion

# region DELETE CATEGORY
@admin_router.callback_query(F.data == 'delete_category')
async def category_operation_type(callback: CallbackQuery, state: FSMContext):
    await state.update_data(sex=None)  # Initialize sex as None to allow skipping
    await callback.answer('')

    # Directly display categories without requiring gender selection
    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.select_categories(
            back_callback='to_admin_main',
            page=1,
            categories_per_page=4,
            language='ru',
            sex=None  # Fetch all categories by default
        )
    )


@admin_router.callback_query(F.data.startswith('selectcategories_'))
async def admin_categories_pagination(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[1])
    data = await state.get_data()
    sex = data.get('sex', None)  # Retrieve selected gender or default to None
    await callback.answer('')

    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.select_categories(
            back_callback='to_admin_main',
            page=page,
            categories_per_page=4,
            language='ru',
            sex=sex
        )
    )


@admin_router.callback_query(F.data.startswith('selectcategory_'))
async def select_category_to_delete(callback: CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split('_')[-1])
    await state.update_data(category_id=category_id)
    await callback.answer('')

    await callback.message.edit_text(
        text=f'Вы выбрали категорию с ID {category_id}.\nВы уверены, что хотите удалить эту категорию? (Y/N)'
    )
    await state.set_state(DeleteCategory.waiting_for_confirmation)


@admin_router.message(DeleteCategory.waiting_for_confirmation)
async def confirm_category_deletion(message: Message, state: FSMContext):
    confirmation = message.text.lower()
    data = await state.get_data()
    category_id = data.get('category_id')

    if confirmation == 'y':
        try:
            # Use async session to delete the category
            deleted = await orm_delete_category(category_id)

            if deleted:
                await message.answer(
                    f'Категория с ID {category_id} была успешно удалена.\nВы в разделе категории.',
                    reply_markup=kb.admin_category
                )
            else:
                await message.answer(
                    f'Категория с ID {category_id} не найдена.',
                    reply_markup=kb.admin_category
                )
        except Exception as e:
            await message.answer(f'Произошла ошибка при удалении категории: {e}')
    elif confirmation == 'n':
        await message.answer('Удаление категории отменено.', reply_markup=kb.admin_category)
    else:
        await message.answer(
            'Неверный ввод. Пожалуйста, введите Y для подтверждения или N для отмены.'
        )
        return

    await state.clear()


# endregion delete
@admin_router.callback_query(F.data == 'to_admin_category')
async def admin_main_back(callback: CallbackQuery):
    await callback.answer('Вы на в пункте категории')
    await callback.message.edit_text('Выберите Действие для Категории', reply_markup=kb.admin_category)


@admin_router.callback_query(F.data == 'to_admin_main')
async def admin_main_back(callback: CallbackQuery):
    await callback.answer('Вы на Главной')
    await callback.message.edit_text('Выберите что будете менять', reply_markup=kb.admin_main)


# endregion

# region ADMIN PRODUCT ADD, CHANGE, DELETE Handlers
# region Add PRODUCT
@admin_router.callback_query(F.data == 'add_product')
async def admin_add_product(callback: CallbackQuery, state: FSMContext):
    await state.update_data(sex=None)  # Initialize `sex` to None to skip gender selection
    await callback.answer('')

    # Display categories directly
    await callback.message.edit_text(
        text='Выберите Категорию Товара',
        reply_markup=await kb.admin_add_product_categories(
            back_callback='to_admin_product',
            page=1,
            categories_per_page=4,
            language='ru',
            sex=None  # Fetch all categories by default
        )
    )


@admin_router.callback_query(F.data.startswith('paddcategories_'))
async def admin_categories_pagination(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[1])
    data = await state.get_data()
    sex = data.get('sex', None)  # Retrieve gender from state or default to None
    await callback.answer('')

    await callback.message.edit_text(
        text='Выберите Категорию Товара.',
        reply_markup=await kb.admin_add_product_categories(
            back_callback='to_admin_product',
            page=page,
            categories_per_page=4,
            language='ru',
            sex=sex
        )
    )


@admin_router.callback_query(F.data.startswith('paddcategory_'))
async def select_category_for_product(callback: CallbackQuery, state: FSMContext):
    category_id = int(callback.data.split('_')[-1])
    await state.update_data(category_id=category_id)
    await callback.answer('')

    # Proceed to the next step
    await callback.message.answer('Введите имя нового товара (RU)', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddProduct.add_name_ru)


@admin_router.message(AddProduct.add_name_ru)
async def admin_add_product_name(message: Message, state=FSMContext):
    await state.update_data(name_ru=message.text)
    await message.answer('Введите Имя Товара одним сообщением(UZ)', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddProduct.add_name_uz)


@admin_router.message(AddProduct.add_name_uz)
async def admin_add_product_name(message: Message, state=FSMContext):
    await state.update_data(name_uz=message.text)
    await message.answer('Введите Описание Товара одним сообщением(RU)', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddProduct.add_description_ru)


@admin_router.message(AddProduct.add_description_ru)
async def admin_add_product_description(message: Message, state=FSMContext):
    await state.update_data(description_ru=message.text)
    await message.answer('Введите Описание Товара одним сообщением(UZ)', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddProduct.add_description_uz)


@admin_router.message(AddProduct.add_description_uz)
async def admin_add_product_description(message: Message, state=FSMContext):
    await state.update_data(description_uz=message.text)
    await message.answer('Введите цену товара только числами(Сум)', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddProduct.add_price)


@admin_router.message(AddProduct.add_price)
async def admin_add_product_price(message: Message, state: FSMContext):
    await state.update_data(price=message.text)
    await message.answer('Отправьте видео товара', reply_markup=ReplyKeyboardRemove())
    await state.set_state(AddProduct.add_video)


@admin_router.message(AddProduct.add_video, F.video)
async def admin_add_product_video(message: Message, state: FSMContext):
    video = message.video.file_id if message.video else None
    await state.update_data(video=video)

    # Save the product with the video
    await save_product_and_clear_state(message, state)


async def save_product_and_clear_state(message: Message, state: FSMContext):
    try:
        data = await state.get_data()
        category_id = data.get('category_id')
        name_ru = data.get('name_ru')
        name_uz = data.get('name_uz')
        description_ru = data.get('description_ru')
        description_uz = data.get('description_uz')
        price = float(data.get('price')) if 'price' in data else None
        video = data.get('video')  # Get the video file_id

        if category_id and name_ru and name_uz and description_ru and description_uz and price and video:
            await orm_add_product(
                name_ru=name_ru,
                name_uz=name_uz,
                description_ru=description_ru,
                description_uz=description_uz,
                price=price,
                category_id=category_id,
                video_url=video  # Pass the video file_id
            )
            await message.answer(
                f'Товар {name_ru} успешно добавлен в категорию №{category_id}',
                reply_markup=kb.admin_product
            )
            await state.clear()
        else:
            await message.answer('Не удалось добавить товар. Пожалуйста, заполните все поля правильно.')

    except Exception as e:
        await message.answer(f'Произошла ошибка при добавлении товара: {str(e)}')


# endregion

#  region CHANGE PRODUCT
@admin_router.callback_query(F.data == 'change_product')
async def admin_main_back(callback: CallbackQuery, state: FSMContext):
    await state.update_data(sex=None)  # Skip gender selection by setting `sex` to None
    await callback.answer('Вы в пункте Продукты')

    await callback.message.edit_text(
        text='Выберите из какой категории продукт',
        reply_markup=await kb.pchcategory_keyboard(
            back_callback='to_admin_product',
            page=1,
            categories_per_page=4,
            language='ru',
            sex=None  # Fetch all categories
        )
    )


@admin_router.callback_query(F.data.startswith('pchcategories_'))
async def categories_pagination(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[1])
    data = await state.get_data()
    sex = data.get('sex', None)  # Retrieve `sex` if it exists; otherwise, it's None
    await callback.answer('')

    await callback.message.edit_text(
        text='Выберите из какой категории продукт',
        reply_markup=await kb.pchcategory_keyboard(
            back_callback='to_admin_product',
            page=page,
            categories_per_page=4,
            language='ru',
            sex=sex  # Pass the value (None to show all categories)
        )
    )


@admin_router.callback_query(F.data.startswith('pchcategory_'))
async def select_category_for_product(callback: CallbackQuery, state: FSMContext):
    data = callback.data.split('_')
    category_id = int(data[1])
    page = int(data[2]) if len(data) == 3 else 1  # Default to page 1 if not provided
    await state.update_data(category_id=category_id)
    await callback.answer('')

    # Load products from the selected category and display them
    products_markup = await kb.products_by_category(category_id, page=1, items_per_row=3)
    await callback.message.edit_text(
        text=f'Выберите продукт (Страница {page})',
        reply_markup=products_markup
    )


@admin_router.callback_query(F.data.startswith('products_'))
async def category_pagination(callback: CallbackQuery, state: FSMContext):
    data = callback.data.split('_')
    category_id = int(data[1])
    page = int(data[2])
    await state.update_data(category_id=category_id)
    await callback.answer('')
    await callback.message.edit_text('Выберите товар',
                                     reply_markup=await kb.products_by_category(category_id, page=page,
                                                                                items_per_row=3))


@admin_router.callback_query(F.data.startswith('product_'))
async def show_product_details(callback: types.CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split('_')[1])
    product = await orm_get_product_by_id(product_id)

    if product:
        await state.update_data(product_id=product_id)
        text = (
            f"Название (RU): {product.name_ru}\n"
            f"Название (UZ): {product.name_uz}\n"
            f"Описание (RU): {product.description_ru}\n"
            f"Описание (UZ): {product.description_uz}\n"
            f"Цена: {product.price} Сум\n\n\n"
            "ВЫБЕРИТЕ ЧТО ХОТИТЕ ИЗМЕНИТЬ"
        )

        await callback.message.delete()

        if product.video_url:
            await callback.message.answer_video(video=product.video_url, caption=text,
                                                reply_markup=kb.admin_product_change)
        else:
            await callback.message.answer(text, reply_markup=kb.admin_product_change)
    else:
        await callback.message.edit_text("Продукт не найден", reply_markup=kb.admin_main)


@admin_router.callback_query(F.data == 'change_p_name')
async def start_change_product_name(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ChangeProduct.ch_name_ru)
    await callback.answer('')
    await callback.message.answer('Введите новое имя продукта', reply_markup=ReplyKeyboardRemove())


@admin_router.message(ChangeProduct.ch_name_ru)
async def set_new_product_name(message: Message, state: FSMContext):
    new_name_ru = message.text
    data = await state.get_data()
    product_id = data.get('product_id')

    updated = await orm_update_product_name_ru(product_id, new_name_ru)

    if updated:
        await message.answer(f'Имя продукта обновлено на {new_name_ru}.\n\nВведите имя на UZ',
                             reply_markup=ReplyKeyboardRemove())
    else:
        await message.answer(f'Продукт с ID {product_id} не найден.', reply_markup=kb.admin_product)
    await state.set_state(ChangeProduct.ch_name_uz)


@admin_router.message(ChangeProduct.ch_name_uz)
async def set_new_product_name(message: Message, state: FSMContext):
    new_name_uz = message.text
    data = await state.get_data()
    product_id = data.get('product_id')

    updated = await orm_update_product_name_uz(product_id, new_name_uz)

    if updated:
        await message.answer(f'Имя продукта обновлено на {new_name_uz}.(UZ)', reply_markup=kb.admin_product)
    else:
        await message.answer(f'Продукт с ID {product_id} не найден.', reply_markup=kb.admin_product)
    await state.clear()


@admin_router.callback_query(F.data == 'change_p_description')
async def start_change_product_description(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ChangeProduct.ch_description_ru)
    await callback.answer('')
    await callback.message.answer('Введите новое описание продукта', reply_markup=ReplyKeyboardRemove())


@admin_router.message(ChangeProduct.ch_description_ru)
async def set_new_product_description(message: Message, state: FSMContext):
    new_description_ru = message.text
    data = await state.get_data()
    product_id = data.get('product_id')

    updated = await orm_update_product_description_ru(product_id, new_description_ru)

    if updated:
        await message.answer(f'Описание продукта обновлено на {new_description_ru}.\n\nДобавь описание на UZ',
                             reply_markup=ReplyKeyboardRemove())
    else:
        await message.answer(f'Продукт с ID {product_id} не найден.', reply_markup=kb.admin_product)
    await state.set_state(ChangeProduct.ch_description_uz)


@admin_router.message(ChangeProduct.ch_description_uz)
async def set_new_product_description(message: Message, state: FSMContext):
    new_description_uz = message.text
    data = await state.get_data()
    product_id = data.get('product_id')

    updated = await orm_update_product_description_uz(product_id, new_description_uz)

    if updated:
        await message.answer(f'Описание продукта обновлено на {new_description_uz}.', reply_markup=kb.admin_product)
    else:
        await message.answer(f'Продукт с ID {product_id} не найден.', reply_markup=kb.admin_product)
    await state.clear()


@admin_router.callback_query(F.data == 'change_p_price')
async def start_change_product_price(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ChangeProduct.ch_price)
    await callback.answer('')
    await callback.message.answer('Введите цену товара в Сумах', reply_markup=ReplyKeyboardRemove())


@admin_router.message(ChangeProduct.ch_price)
async def set_new_product_price(message: Message, state: FSMContext):
    new_price = int(message.text)
    data = await state.get_data()
    product_id = data.get('product_id')

    updated = await orm_update_product_price(product_id, new_price)

    if updated:
        await message.answer(f'Новая цена товара {new_price}.', reply_markup=kb.admin_product)
    else:
        await message.answer(f'Продукт с ID {product_id} не найден.', reply_markup=kb.admin_product)
    await state.clear()


@admin_router.callback_query(F.data == 'change_p_video')
async def start_change_product_video(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ChangeProduct.ch_video)
    await callback.answer('')
    await callback.message.answer('Отправьте новое видео продукта', reply_markup=ReplyKeyboardRemove())


@admin_router.message(ChangeProduct.ch_video, F.video)
async def set_new_product_video(message: Message, state: FSMContext):
    new_video = message.video.file_id
    data = await state.get_data()
    product_id = data.get('product_id')
    updated = await orm_update_product_video(product_id, new_video)

    if updated:
        await message.answer(f'Видео продукта обновлено.', reply_markup=kb.admin_product)
    else:
        await message.answer(f'Продукт с ID {product_id} не найден.', reply_markup=kb.admin_product)
    await state.clear()



# endregion

# region Delete Product

@admin_router.callback_query(F.data == 'delete_product')
async def admin_main_back(callback: CallbackQuery, state: FSMContext):
    await state.update_data(sex=None)  # Skip gender selection by setting `sex` to None
    await callback.answer('Удаление')

    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.pdcategory_keyboard(
            back_callback='to_admin_product',
            page=1,
            categories_per_page=4,
            language='ru',
            sex=None  # Fetch all categories without filtering by gender
        )
    )


@admin_router.callback_query(F.data.startswith('pdcategories_'))
async def categories_pagination(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[1])
    data = await state.get_data()
    sex = data.get('sex', None)  # Default to None (no gender filtering)
    language_code = 'ru'
    await callback.answer('')

    await callback.message.edit_text(
        text='Выберите категорию.',
        reply_markup=await kb.pdcategory_keyboard(
            back_callback='to_admin_product',
            page=page,
            categories_per_page=4,
            language=language_code,
            sex=sex  # Pass the value (None to show all categories)
        )
    )


@admin_router.callback_query(F.data.startswith('pdcategorie_'))
async def admin_select_category_for_product(callback: CallbackQuery, state: FSMContext):
    data = callback.data.split('_')
    category_id = int(data[1])
    page = int(data[2]) if len(data) == 3 else 1  # Default to page 1 if no page is provided
    await state.update_data(category_id=category_id)
    await callback.answer('')

    # Load products from the selected category and display them
    products_markup = await kb.products_to_delete(category_id, page, items_per_row=3)
    await callback.message.edit_text(
        text=f'Выберите продукт для удаления (Страница {page})',
        reply_markup=products_markup
    )


@admin_router.callback_query(F.data.startswith('dproducts_'))
async def admin_category_pagination(callback: CallbackQuery, state: FSMContext):
    data = callback.data.split('_')
    category_id = int(data[1])
    page = int(data[2])
    await state.update_data(category_id=category_id)
    await callback.answer('')

    await callback.message.edit_text(
        text='Выберите товар',
        reply_markup=await kb.products_to_delete(category_id, page)
    )


@admin_router.callback_query(F.data.startswith('dproduct_'))
async def show_product_details(callback: CallbackQuery, state: FSMContext):
    product_id = int(callback.data.split('_')[1])
    product = await orm_get_product_by_id(product_id)

    if product:
        await state.update_data(product_id=product_id,
                                product_name=product.name_ru)  # Store product_id and name in state
        text = (
            f"Название: {product.name_uz}\n"
            f"Описание: {product.description_ru}\n"
            f"Цена: {product.price} Сум\n\n\n"
            "Вы уверены что хотите удалить этот продукт? (Y/N)"
        )
        if product.image_url:  # Using the image_url for the photo
            await callback.message.delete()  # Delete the previous message
            await callback.message.answer_photo(photo=product.image_url, caption=text, reply_markup=None)
        else:
            await callback.message.edit_text(text, reply_markup=None)

        await state.set_state(DeleteProduct.delete_product)
    else:
        await callback.message.edit_text("Продукт не найден", reply_markup=kb.admin_main)


@admin_router.message(DeleteProduct.delete_product)
async def confirm_product_deletion(message: Message, state: FSMContext):
    confirmation = message.text.strip().lower()
    data = await state.get_data()
    product_id = data.get('product_id')
    product_name = data.get('product_name')

    if confirmation == 'y':

        deleted = await orm_delete_product_by_id(product_id)

        if deleted:
            await message.answer(f'Продукт "{product_name}" успешно удален.', reply_markup=kb.admin_product)
        else:
            await message.answer(f'Ошибка при удалении продукта "{product_name}".')
    elif confirmation == 'n':
        await message.answer('Удаление отменено.')
    else:
        await message.answer('Пожалуйста, введите "Y" для подтверждения удаления или "N" для отмены.')

    # Clear the state after handling the confirmation
    await state.clear()


# endregion

@admin_router.callback_query(F.data == 'to_admin_product')
async def admin_main_back(callback: CallbackQuery):
    await callback.message.delete()
    await callback.message.answer('Выберите Действие для Товара', reply_markup=kb.admin_product)


@admin_router.callback_query(F.data == 'newsletter')
async def admin_main_back(callback: CallbackQuery):
    await callback.message.delete()
    await callback.message.answer('Напишите /newsletter в начале вашего сообщения для рассылки всем пользователям бота')


@admin_router.message(Command("newsletter"))
async def broadcast_message(message: Message, bot: Bot):
    # Get all user IDs from the database
    user_ids = await orm_get_all_user_ids()
    # Get the text to broadcast from the message
    text_to_broadcast = message.text[len('/broadcast '):]

    # Send the message to all users
    for user_id in user_ids:
        try:
            await bot.send_message(user_id, text_to_broadcast)
            logging.info(f"Сообщение отправлено пользователю {user_id}")
        except Exception as e:
            logging.error(f"Ошибка при отправке сообщения пользователю {user_id}: {e}")

    await message.reply("Сообщение успешно отправлено всем пользователям.")


# endregion


@admin_router.message(Command('getallorders'))
async def send_all_orders(message: types.Message, session: AsyncSession):
    orders = await orm_get_all_excel_orders()  # Ensure it retrieves all necessary fields

    if not orders:
        await message.reply("No orders found.")
        return

    data = {
        "Order ID": [],
        "Category Name (RU)": [],
        "Product Name (RU)": [],
        "Product Quantity": [],
        "Initial Cost": [],
        "Promo Code": [],
        "Discount %": [],
        "Total Cost": [],
        "Customer Name": [],
        "Username": [],
        "Phone Number": [],
        "Orders Time": [],
        "Location": [],
        "Bonus Product Name": [],  # ✅ Added
        "Location Name": [],  # ✅ Added
        "Status": []
    }

    for order in orders:
        data["Order ID"].append(order.get("order_id"))
        data["Category Name (RU)"].append(order.get("category_name_ru", "N/A"))
        data["Product Name (RU)"].append(order.get("product_name_ru", "N/A"))
        data["Product Quantity"].append(order.get("product_quantity", 0))
        data["Initial Cost"].append(order.get("initial_cost", 0.0))
        data["Promo Code"].append(order.get("promo_code_name", "N/A"))
        data["Discount %"].append(order.get("promo_discount_percentage", 0.0))
        data["Total Cost"].append(order.get("total_cost", 0.0))
        data["Customer Name"].append(order.get("customer_name", "N/A"))
        data["Username"].append(order.get("username", "N/A"))
        data["Phone Number"].append(order.get("phone_number", "N/A"))
        data["Orders Time"].append(order.get("order_created_at", "N/A"))
        data["Bonus Product Name"].append(order.get("bonus_product_name", "N/A"))  # ✅ Added
        data["Location Name"].append(order.get("location_name", "N/A"))  # ✅ Added

        location = f"{order.get('latitude', 'N/A')}, {order.get('longitude', 'N/A')}"
        data["Location"].append(location)

        data["Status"].append(order.get("status", "pending"))

    df = pd.DataFrame(data)

    # Create a temporary file to save the Excel file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        file_path = tmp.name
        df.to_excel(file_path, index=False)

    # Open file to apply colors
    wb = load_workbook(file_path)
    ws = wb.active

    # Define colors
    status_colors = {
        "FINISHED": "00FF00",  # Green
        "PENDING": "ADD8E6",  # Light Blue
        "CANCELLED": "FF0000"  # Red
    }

    # Find "Status" column index
    status_col_index = df.columns.get_loc("Status") + 1

    for row in range(2, len(df) + 2):  # Start from row 2 (skip headers)
        cell = ws.cell(row=row, column=status_col_index)

        status_value = str(cell.value).strip().upper()  # Normalize status text

        fill_color = status_colors.get(status_value, "FFFFFF")  # Default to white
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Save the modified Excel file
    wb.save(file_path)
    wb.close()

    # Send file
    file_input = FSInputFile(file_path)
    await message.reply_document(file_input)

    # Remove file after sending
    os.remove(file_path)


@admin_router.message(Command("ExcelOrdersClear"))
async def broadcast_message(message: Message):
    try:
        await orm_delete_all_excel_orders()
        await message.answer("Все записи с таблицы сведений удалены успешно!.")
    except Exception as e:
        await message.answer(f"Failed to clear orders: {str(e)}")


@admin_router.message(Command("OrderCancelation"))
async def cancel_order(message: Message):
    args = message.text.split()[1:]  # Extract arguments after the command

    if not args:
        await message.answer("Пожалуйста, укажите ID заказа, который хотите отменить.")
        return

    try:
        order_id = int(args[0])  # Extract and convert order ID
    except ValueError:
        await message.answer("Неверный формат ID заказа. Введите числовой идентификатор.")
        return

    user = await orm_get_user_by_tg_id(message.from_user.id)
    if not user:
        await message.answer("Вы не зарегистрированы в системе.")
        return

    # Check if the order exists and belongs to the user
    order = await orm_get_order_by_id(order_id)
    if not order or order.user_id != user.id:
        await message.answer("Заказ не найден или вы не являетесь его владельцем.")
        return

    # Delete the order
    await orm_delete_order_by_id(order_id)

    # Update the status of the relevant ExcelOrders to 'cancelled'
    await update_excel_order_status_to_cancelled(order_id)

    await message.answer(f"Заказ {order_id} был успешно отменен.")


@admin_router.message(Command("OrderFinish"))
async def finish_order(message: Message):
    args = message.text.split()[1:]  # Extract arguments after the command

    if not args:
        await message.answer("Пожалуйста, укажите ID заказа, который хотите завершить.")
        return

    try:
        order_id = int(args[0])  # Extract and convert order ID
    except ValueError:
        await message.answer("Неверный формат ID заказа. Введите числовой идентификатор.")
        return

    user = await orm_get_user_by_tg_id(message.from_user.id)
    if not user:
        await message.answer("Вы не зарегистрированы в системе.")
        return

    # Check if the order exists and belongs to the user
    order = await orm_get_order_by_id(order_id)
    if not order or order.user_id != user.id:
        await message.answer("Заказ не найден или вы не являетесь его владельцем.")
        return

    # Update the status of the order to "finished"
    await update_order_status_to_finished(order_id)

    await message.answer(f"Заказ {order_id} был успешно завершен.")


@admin_router.message(Command("GetLocation"))
async def get_location(message: Message):
    args = message.text.split()[1:]  # Extract arguments after the command

    if not args:
        await message.answer("Пожалуйста, укажите ID заказа, чтобы получить местоположение.")
        return

    try:
        order_id = int(args[0])  # Convert order ID to integer
    except ValueError:
        await message.answer("Неверный формат ID заказа. Введите числовой идентификатор.")
        return

    # Retrieve order from database
    order = await orm_get_order_by_id(order_id)
    if not order:
        await message.answer("Заказ не найден.")
        return

    # Retrieve user associated with the order
    user = order.user
    if not user:
        await message.answer("Пользователь, связанный с этим заказом, не найден.")
        return

    # Extract location
    latitude = user.latitude
    longitude = user.longitude

    if latitude is None or longitude is None:
        await message.answer(f"У пользователя, оформившего заказ {order_id}, нет сохраненного местоположения.")
        return

    # Send the order ID message + location to the group
    for group in GROUP_CHAT_IDS_WITH_THREADS:
        chat_id = group["chat_location"]
        thread_id = group["message_thread_id"]

        # Send a message with the Order ID before sending the location
        await message.bot.send_message(
            chat_id=chat_id,
            text=f"📍 Местоположение для заказа #{order_id}",
            message_thread_id=thread_id
        )

        # Send the location
        await message.bot.send_location(
            chat_id=chat_id,
            latitude=float(latitude),
            longitude=float(longitude),
            message_thread_id=thread_id
        )

    await message.answer(f"Местоположение заказа {hbold(order_id)} отправлено в группу.")


@admin_router.message(Command("add_promocode"))
async def start_add_promocode(message: Message, state: FSMContext):
    await message.answer("Введите текст промокода:")
    await state.set_state(AddPromoCode.enter_promo_code)


@admin_router.message(AddPromoCode.enter_promo_code)
async def process_promo_code_text(message: Message, state: FSMContext):
    promo_code = message.text.strip()

    # Validate the promo code format
    if not promo_code.isalnum():
        await message.answer("Промокод должен содержать только буквы и цифры.")
        return

    # Check if the promo code already exists
    if await orm_promocode_exists(promo_code):
        await message.answer("Промокод с таким текстом уже существует.")
        return

    await state.update_data(promo_code=promo_code)
    await message.answer("Введите ID товара, для которого действует промокод (или 'все' для всех товаров):")
    await state.set_state(AddPromoCode.enter_product_id)


@admin_router.message(AddPromoCode.enter_product_id)
async def process_product_id(message: Message, state: FSMContext):
    user_input = message.text.strip()

    if user_input.lower() == "все":
        await state.update_data(product_id=None, is_global=True)
        await message.answer("Введите размер скидки в процентах (например, 10 для 10%):")
        await state.set_state(AddPromoCode.enter_discount)
    else:
        try:
            product_id = int(user_input)  # Convert the product ID to an integer

            # Check if the product exists
            if not await orm_product_exists(product_id):
                await message.answer(f"Товар с ID {product_id} не найден.")
                return

            await state.update_data(product_id=product_id, is_global=False)
            await message.answer("Введите размер скидки в процентах (например, 10 для 10%):")
            await state.set_state(AddPromoCode.enter_discount)
        except ValueError:
            await message.answer("Неверный формат ID товара. Введите числовой идентификатор или 'все'.")


@admin_router.message(AddPromoCode.enter_discount)
async def process_discount(message: Message, state: FSMContext):
    try:
        discount = float(message.text)  # Convert the discount to a float
        if discount <= 0 or discount > 100:
            await message.answer("Скидка должна быть больше 0% и не больше 100%.")
            return

        await state.update_data(discount=discount)
        await message.answer("Введите срок действия промокода в формате ГГГГ-ММ-ДД (или 'нет' для отсутствия срока):")
        await state.set_state(AddPromoCode.enter_expiry_date)
    except ValueError:
        await message.answer("Неверный формат скидки. Введите число (например, 10 для 10%).")


@admin_router.message(AddPromoCode.enter_expiry_date)
async def process_expiry_date(message: Message, state: FSMContext):
    user_input = message.text.strip()

    if user_input.lower() == "нет":
        expiry_date = None
    else:
        try:
            expiry_date = datetime.strptime(user_input, "%Y-%m-%d")
        except ValueError:
            await message.answer("Неверный формат даты. Введите дату в формате ГГГГ-ММ-ДД или 'нет'.")
            return

    # Get the data from the state
    data = await state.get_data()
    promo_code = data.get("promo_code")
    product_id = data.get("product_id")
    discount = data.get("discount")
    is_global = data.get("is_global", False)

    # Save the promo code in the database
    await orm_add_promocode(promo_code, product_id, discount, is_global, expiry_date)

    await message.answer(f"Промокод '{promo_code}' успешно добавлен.")
    await state.clear()  # Clear the state


@admin_router.message(Command("promocodedelete"))
async def delete_promocode(message: Message):
    args = message.text.split()[1:]  # Extract arguments after the command

    if not args:
        await message.answer("Пожалуйста, укажите ID промокода для удаления.")
        return

    try:
        promo_code_id = int(args[0])  # Convert the promo code ID to an integer
    except ValueError:
        await message.answer("Неверный формат ID промокода. Введите числовой идентификатор.")
        return

    # Delete the promo code
    if await orm_delete_promocode(promo_code_id):
        await message.answer(f"Промокод с ID {promo_code_id} успешно удален.")
    else:
        await message.answer(f"Промокод с ID {promo_code_id} не найден.")


@admin_router.message(Command("add_bonus"))
async def add_bonus_start(message: Message, state: FSMContext):
    await message.answer("Введите название бонусного продукта на русском:")
    await state.set_state(AddBonusProduct.name_ru)


# Prompt for name in Russian
@admin_router.message(AddBonusProduct.name_ru)
async def process_bonus_name_ru(message: Message, state: FSMContext):
    name_ru = message.text.strip()

    if not name_ru:
        await message.answer("Название не может быть пустым. Введите название бонусного продукта на русском:")
        return

    await state.update_data(name_ru=name_ru)
    await message.answer("Введите название бонусного продукта на узбекском:")
    await state.set_state(AddBonusProduct.name_uz)


# Prompt for name in Uzbek
@admin_router.message(AddBonusProduct.name_uz)
async def process_bonus_name_uz(message: Message, state: FSMContext):
    name_uz = message.text.strip()

    if not name_uz:
        await message.answer("Название не может быть пустым. Введите название бонусного продукта на узбекском:")
        return

    await state.update_data(name_uz=name_uz)
    await message.answer("Введите описание бонусного продукта на русском:")
    await state.set_state(AddBonusProduct.description_ru)


# Prompt for description in Russian
@admin_router.message(AddBonusProduct.description_ru)
async def process_bonus_description_ru(message: Message, state: FSMContext):
    description_ru = message.text.strip()

    if not description_ru:
        await message.answer("Описание не может быть пустым. Введите описание бонусного продукта на русском:")
        return

    await state.update_data(description_ru=description_ru)
    await message.answer("Введите описание бонусного продукта на узбекском:")
    await state.set_state(AddBonusProduct.description_uz)


# Prompt for description in Uzbek
@admin_router.message(AddBonusProduct.description_uz)
async def process_bonus_description_uz(message: Message, state: FSMContext):
    description_uz = message.text.strip()

    if not description_uz:
        await message.answer("Описание не может быть пустым. Введите описание бонусного продукта на узбекском:")
        return

    await state.update_data(description_uz=description_uz)
    await message.answer("Отправьте URL изображения бонусного продукта (или напишите 'нет'):")
    await state.set_state(AddBonusProduct.image_url)


@admin_router.message(AddBonusProduct.image_url)
async def process_bonus_image(message: Message, state: FSMContext):
    image_url = message.text.strip()

    if image_url.lower() == "нет":
        image_url = None

    data = await state.get_data()
    name_ru = data["name_ru"]
    name_uz = data["name_uz"]
    description_ru = data["description_ru"]
    description_uz = data["description_uz"]

    await message.answer("Введите количество рефералов, необходимых для получения бонуса:")
    await state.update_data(image_url=image_url)
    await state.set_state(AddBonusProduct.required_referrals)


@admin_router.message(AddBonusProduct.required_referrals)
async def process_bonus_required_referrals(message: Message, state: FSMContext):
    try:
        required_referrals = int(message.text.strip())
    except ValueError:
        await message.answer("Пожалуйста, введите корректное число для количества рефералов.")
        return

    # Ensure we have valid data from previous steps
    data = await state.get_data()
    name_ru = data["name_ru"]
    name_uz = data["name_uz"]
    description_ru = data["description_ru"]
    description_uz = data["description_uz"]
    image_url = data.get("image_url")

    # Save the bonus product to the database
    await orm_add_bonus_product(name_ru, name_uz, description_ru, description_uz, image_url, required_referrals)

    await message.answer(f"Бонусный продукт '{name_ru}' успешно добавлен!")
    await state.clear()


@admin_router.message(Command("delete_bonus"))
async def delete_bonus(message: Message):
    args = message.text.split()[1:]  # Extract arguments after the command

    if not args:
        await message.answer("Пожалуйста, укажите ID бонусного продукта для удаления.")
        return

    try:
        bonus_id = int(args[0])  # Convert the bonus product ID to an integer
    except ValueError:
        await message.answer("Неверный формат ID бонусного продукта. Введите числовой идентификатор.")
        return

    # Delete the bonus product
    if await orm_delete_bonus_product(bonus_id):
        await message.answer(f"Бонусный продукт с ID {bonus_id} успешно удален.")
    else:
        await message.answer(f"Бонусный продукт с ID {bonus_id} не найден.")
